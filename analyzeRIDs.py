from openpyxl import load_workbook
import openpyxl

class RawData():
#object for each RID and Rterms Excel file
    def __init__(self, fN, newfN):
        self.fN = fN
        self.newfN = newfN
        self.workbook = load_workbook(filename=fN)
        self.sheet = self.workbook.active
        self.rDict = {}
        self.sortedKeys = []

    def dictOfterms(self):
    #create dictionary where key is the RadLex item, values are the filepaths
        for row in self.sheet.iter_rows(values_only=True):
            filepath, radlex = row
            rs = radlex.split(", ")
            for r in rs:
                if r in self.rDict:
                    self.rDict[r].append(filepath)
                else:
                    self.rDict[r] = [filepath]
    
    def sortByLen(self):
    #sort by number
        self.sortedKeys = sorted(self.rDict, key=lambda k: len(self.rDict[k]), reverse=True)
    
    def writeResults(self):
    #frequencies of all RadLex items sorted in order of most to least frequent
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        row = 1
        for k in self.sortedKeys:
            sheet.cell(row=row, column=1, value=k)
            sheet.cell(row=row, column=2, value=len(self.rDict[k]))
            sheet.cell(row=row, column=3, value=', '.join(self.rDict[k]))
            row += 1

        workbook.save(filename=self.newfN)
    
    def analyze(self):
    #call all the necessary methods in order
        self.dictOfterms()
        self.sortByLen()
        self.writeResults()
    

IDs = RawData("FileRIDs.xlsx", "freqRIDs.xlsx")
IDs.analyze()
Rterms = RawData("FileRterms.xlsx", "freqRterms.xlsx")
Rterms.analyze()
