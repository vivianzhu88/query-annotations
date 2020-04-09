import urllib.request, urllib.error, urllib.parse
import requests
import json
import os
import openpyxl
from pprint import pprint

REST_URL = "http://data.bioontology.org"
ONT = "&ontologies=RADLEX"
API_KEY = ""
    
def get_annotations(text, url):
    headers = {'Authorization': 'apikey token=' + API_KEY}
    data = text

    response = requests.request("POST",url,headers=headers,data=data)
    print(response.text)
    return json.loads(response.text)
    
def mapRIDs(text, get_class=True):
#take annotations and map to Excel spreadsheet of labels and corresponding RIDs
    annotations = get_annotations(text, REST_URL + "/annotator?text=" + urllib.parse.quote(text) + ONT)
    
    RIDs = []
    #iterate through annotations
    for result in annotations:
        class_details = result["annotatedClass"]
        if get_class:
            try:
                class_details = get_json(result["annotatedClass"]["links"]["self"])
            except urllib.error.HTTPError:
                print(f"Error retrieving {result['annotatedClass']['@id']}")
                continue
        
        #get each RIDs correlating with text
        id = class_details["@id"]
        RIDs.append(id[22:])
    return RIDs

class File():
    def __init__(self, filename):
    #initialize each file
        self.filename = filename
        self.RIDs = []
        self.text = ""
    
    def openFile(self):
    #open file and retrieve text
    #has .ore .txt .lsx .ocx .csv .xls .son .xml files
        with open(self.filename, 'rb') as file:
            self.text = file.read()
    
    def getName(self):
        return self.filename
    
    def getRIDs(self):
        self.RIDs = mapRIDs(self.text)
        return self.RIDs

def toSpreadsheet(RIDs):
    #put data into to Excel spreadsheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    row = 1
    for key,values in annDict.items():
        sheet.cell(row=row, column=1, value=key)
        sheet.cell(row=row, column=2, value=', '.join(values))
        row += 1

    workbook.save(filename="fileRIDs.xlsx")

#put all of file paths in Chest_and_Lung_Collections directory into a list
filesList = []
RIDsDict = {}

dir_path = "/Users/vivianzhu/Documents/Annotator/Chest_and_Lung_Collections/"
for dp,_,filenames in os.walk(dir_path):
   for f in filenames:
       f = File(os.path.abspath(os.path.join(dp, f))[len(dir_path):])
       filesList.append(f)

for f in filesList:
    f.openFile()
    RIDsDict[f.getName()] = f.getRIDs()

toSpreadsheet(RIDsDict)
    
    

