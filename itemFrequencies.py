from openpyxl import load_workbook
import openpyxl
import json

class RawData():
#object for each RID and Rterms Excel file
    def __init__(self, fN, itemName):
        self.fN = fN
        self.itemName = itemName
        self.workbook = load_workbook(filename=fN)
        self.sheet = self.workbook.active
        self.rDict = {}
        self.sortedKeys = []
    
    def getItemName(self):
    #return item name
        return self.itemName
    
    def getItems(self):
    #return list of RadLex RIDs or Rterms
        items = []
        for k in self.sortedKeys:
            items.append(k)
        return items

    def dictOfterms(self):
    #create dictionary where key is the RadLex item, values are the filepaths
        for row in self.sheet.iter_rows(min_row=2, values_only=True):
            filepath, radlex = row
            rs = radlex.split(" | ")
            for r in rs:
                if r in self.rDict.keys():
                    self.rDict[r].append(filepath)
                else:
                    self.rDict[r] = [filepath]
    
    def sortByLen(self):
    #sort by number
        self.sortedKeys = sorted(self.rDict, key=lambda k: len(self.rDict[k]), reverse=True)
    
    def writeResults(self, rawdata):
    #frequencies of all RadLex items sorted in order of most to least frequent
        
        output = []
        rdItems = rawdata.getItems()
        c = 0
        for k in self.sortedKeys:
            d = {}
            d[self.itemName] = k
            d[rawdata.getItemName()] = rdItems[c]
            d["Count"] = len(self.rDict[k])
            d["File Paths"] = ' | '.join(self.rDict[k])
            output.append(d)
            c += 1
        
        with open("freq.json", "w") as outfile:
            json.dump(output, outfile)
    
    def analyze(self):
    #call all the necessary methods in order
        self.dictOfterms()
        self.sortByLen()
    

IDs = RawData("filepath_and_RIDs.xlsx", "RIDs")
IDs.analyze()
Rterms = RawData("filepath_and_Rterms.xlsx", "Rterms")
Rterms.analyze()
IDs.writeResults(Rterms)
