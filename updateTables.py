from openpyxl import load_workbook
import openpyxl
import json
import pandas as pd
import numpy as np
from csv import writer
from csv import reader

class RawData():
#object for each RID and Rterms Excel file
    def __init__(self, fN):
        self.fN = fN
        self.raw_data = pd.read_csv(fN)
        self.IDs_and_terms = {}

    def getIDTermData(self):
    #create dictionary where key is the ID, values are the terms
        ID_files = ["filepath_and_SIDs.xlsx", "filepath_and_RIDs.xlsx"]
        term_files = ["filepath_and_Sterms.xlsx", "filepath_and_Rterms.xlsx"]
        
        IDs = []
        terms = []
        for file in ID_files:
            workbook = load_workbook(filename=file)
            sheet = workbook.active
            for row in sheet.iter_rows(min_row=2, values_only=True):
                filepath,ID = row
                IDs+=ID.split(" | ")
        for file in term_files:
            workbook = load_workbook(filename=file)
            sheet = workbook.active
            for row in sheet.iter_rows(min_row=2, values_only=True):
                filepath,term = row
                terms+=term.split(" | ")
        
        for i in range(len(IDs)):
            if terms[i].lower() not in self.IDs_and_terms.keys():
                self.IDs_and_terms[terms[i].lower()] = [IDs[i]]
            elif IDs[i] not in self.IDs_and_terms[terms[i].lower()]:
                self.IDs_and_terms[terms[i].lower()] += [IDs[i]]
    
    def findReplace(self):
    #find SRT terms and replace with RADLEX
        ID_column1 = list(self.raw_data["Finding.CodeValue"])
        term_column1 = list(self.raw_data["Finding.CodeMeaning"])
        ont_column1 = list(self.raw_data["Finding.CodingSchemeDesignator"])
        ID_column2 = list(self.raw_data["Finding Site.CodeValue"])
        term_column2 = list(self.raw_data["Finding Site.CodeMeaning"])
        ont_column2 = list(self.raw_data["Finding Site.CodingSchemeDesignator"])
        
        for c in range (1, len(term_column1[1:])):
            tcol = str(term_column1[c])
            tcol = tcol.lower()
            icol = str(ID_column1[c])
            if (tcol in self.IDs_and_terms.keys()) and ("RID" not in icol):
                if len(self.IDs_and_terms[tcol]) > 1:
                    for i in self.IDs_and_terms[tcol]:
                        if "RID" in i:
                            ID_column1[c] = i #Replace SNOMED ID with RADLEX ID if multiple IDs in a term
                            ont_column1[c] = "RADLEX"
                            print("replaced",icol,"with",i)
                            break
                else:
                    ID_column1[c] = self.IDs_and_terms[tcol][0] #Replace SNOMED ID with RADLEX ID if 1 ID in a term
        
        for c in range (1, len(term_column2[1:])):
            tcol = str(term_column2[c])
            tcol = tcol.lower()
            icol = str(ID_column2[c])
            if (tcol in self.IDs_and_terms.keys()) and ("RID" not in icol):
                if len(self.IDs_and_terms[tcol]) > 1:
                    for i in self.IDs_and_terms[tcol]:
                        if "RID" in i:
                            ID_column2[c] = i #Replace SNOMED ID with RADLEX ID if multiple IDs in a term
                            ont_column2[c] = "RADLEX"
                            print("replaced",icol,"with",i)
                            break
                else:
                    ID_column2[c] = self.IDs_and_terms[tcol][0] #Replace SNOMED ID with RADLEX ID if 1 ID in a term

        self.raw_data["Finding.CodeValue"] = ID_column1
        self.raw_data["Finding.CodingSchemeDesignator"] = ont_column1
        self.raw_data["Finding Site.CodeValue"] = ID_column2
        self.raw_data["Finding Site.CodingSchemeDesignator"] = ont_column2
        self.raw_data.to_csv(self.fN[40:])
    
    def findAdd(self):
    #find term and add RADLEX
        add_column = []
        term_column = list(self.raw_data["anatomy"])
        for col in term_column:
            col = col.lower()
            if col in self.IDs_and_terms.keys():
                if len(self.IDs_and_terms[col]) > 1:
                    for i in self.IDs_and_terms[col]:
                        if "RID" in i:
                            print("from", col, "added", i)
                            add_column.append(i) #Add RADLEX ID if multiple IDs in a term
                            break
                else:
                    print("from", col, "added", self.IDs_and_terms[col][0])
                    add_column.append(self.IDs_and_terms[col][0]) #Add whatever ID is available
            else:
                add_column.append("") #Add blank if no matching ID
        
        self.raw_data.insert(1, "RADLEX/SNOMED ID", add_column)
        self.raw_data.to_csv(self.fN[40:])
    
    def replacement(self):
    #call all the necessary methods in order
        self.getIDTermData()
        if "2018" in self.fN:
            self.findReplace()
        elif "2017" in self.fN:
            self.findAdd()
    

the_2018 = RawData("Chest_and_Lung_Collections/CCC_RSNA2018/CrowdsCureCancer2018-Results.csv")
the_2018.replacement()
the_2017 = RawData("Chest_and_Lung_Collections/CCC_RSNA2017/CrowdsCureCancer2017Annotations.csv")
the_2017.replacement()
